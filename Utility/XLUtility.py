import openpyxl


def NoRows(path,sheetname):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetname]
    rows = sheet.max_row
    return rows


def readdata(path,sheetname,rowno,colno):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetname]
    value = sheet.cell(row=rowno,column=colno).value
    return value

def writedata(path,sheetname,rowno,colno,data):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetname]
    value = sheet.cell(row=rowno,column=colno,value=data)
    workbook.save(path)



